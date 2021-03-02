#!/usr/bin/env python2
#-*- coding:utf-8 -*-

import openpyxl 
from openpyxl import load_workbook    #读取excel文件
from openpyxl import Workbook         # 创建xlsx文件

'''
根据下机单提取加测样本下机单，并生成samplelist文件
'''

class ExtractPath(object):

    def __init__(self,args):
        self.args = args
        self.infile = args['infile']
        self.people = args['people']
        self.number = args['number']
        self.targetsam = args['targetsam']
        self.result = args['result']
        self.sampleinfo = args['sampleinfo']
        self.sam_lists = []
        self.sample_info_dict = self.get_sample_info_dict() if self.sampleinfo else {}
        
    def x2t(self):
        '''将下机单进行格式转换
        xlsx转换为文本文件,只保留需要的行文件
        '''
        wb = load_workbook(self.infile)
        ws = wb.active
        out = open(self.infile.replace('.xlsx', '.xls'), 'w')
        for row in  ws.rows:
            n = 0
            cmd = ''
            for cell in row:
                n += 1
                value = cell.value
                if isinstance(value, long):
                    value = str(value)
                else:
                    value = value.encode('utf-8')
                
                if n in [2,6,8,10,21,22,25,29,30]:
                    cmd += value + '\t'
            out.write("{cmd}".format(**locals()))
            out.write("\n")

    def show_people_info(self):
        '''展示指定信息人员的加测分期信息
        '''
        info_dict = {}
        with open(self.infile.replace('.xlsx','.xls')) as fr:
            for line in fr:
                linelist = line.strip().split('\t')
                if linelist[0].lower().startswith('lane'):
                    continue
                sam = linelist[4]
                fenqi = linelist[5]
                if sam in info_dict:
                    if fenqi in info_dict[sam]:
                        pass
                    else:
                        info_dict[sam].append(fenqi)
                else:
                    info_dict[sam] = [fenqi]

        for sam in info_dict:
            if self.people in sam:
                print("\n".join(info_dict[sam]))

    def get_target_sam_lists(self):
        if self.targetsam:
            with open(self.targetsam) as fr:
                for line in fr:
                    sam = line.strip()
                    self.sam_lists.append(sam)
        if self.sam_lists:
            print('\033[1;32m提取指定样本的下机单: {}\033[0m'.format("|".join(self.sam_lists)))

    def change_str(self,item):
        '''读取excel表格时，汉字和字母的方式不一样
        '''
        if isinstance(item, long):
            return str(item)
        else:
            item = item.encode('utf-8')
            return item

    def get_sample_info_dict(self):
        '''下机单中的样本名称，可能不是标准名称,
        需要通过sampleinfo文件进行校正
        sample_info_dict 
        {
            'novoid': [送样名称,标准名称]
        }
        '''
        sample_info_dict = {}
        wb = load_workbook(self.sampleinfo)
        ws = wb.active

        n = 16 # sampleinfo从15行记录信息
        while ws['L{}'.format(n)].value:
            name = ws['L{}'.format(n)].value
            standername = ws['M{}'.format(n)].value
            novoid = ws['N{}'.format(n)].value

            n += 1
            name = self.change_str(name)
            standername = self.change_str(standername)
            novoid = self.change_str(novoid)

            sample_info_dict[novoid] = [str(name),str(standername)]

        return sample_info_dict

    def get_stander_name(self,samid,novoid):
        '''利用sample_info，得到每个样本的标准名称
        根据下机单中的novoid,对样本名称进行校正
        '''
        standername = samid
        if self.sample_info_dict:
            if novoid in self.sample_info_dict:
                if samid == self.sample_info_dict[novoid][0]:
                    standername = self.sample_info_dict[novoid][1]
                else:
                    standername = 'warning'
            else:
                standername = 'wrong'

        return standername

    def extract_target_path(self):
        '''提取指定分期编号的sample_list文件
        novoid一样，但是samid不一样，警告提醒
        '''
        warn_novoid = []
        wrong_novoid = []
        with open(self.infile.replace('.xlsx', '.xls')) as fr, \
            open(self.result,'w') as fw:
            for line in fr:
                linelist = line.strip().split('\t')
                if linelist[0].lower().startswith('lane'):
                    fw.write('#Ori_Lane\tPatientID\tSampleID\tLibID\tNovoID\tIndex\tPath\n')
                    continue              
                lane = linelist[0]
                samid = linelist[7]
                libid = linelist[8]
                novoid = linelist[6]
                index = linelist[1]
                path = linelist[2]
                fenqi = linelist[5]
                
                # 提取指定分期路径
                if fenqi == self.number:
                    standername = self.get_stander_name(samid,novoid) 
                    if standername == 'warning':
                        warn_novoid.append((novoid,samid,self.sample_info_dict[novoid][0]))
                    if standername == 'wrong':
                        wrong_novoid.append((novoid,samid)) 
                    
                    if self.targetsam:
                        if standername in self.sam_lists:
                            newline = '{lane}\t{standername}\t{standername}\t{libid}\t{novoid}\t{index}\t{path}\n'.format(**locals())
                            fw.write(newline)
                    else:
                        newline = '{lane}\t{standername}\t{standername}\t{libid}\t{novoid}\t{index}\t{path}\n'.format(**locals())
                        fw.write(newline)

        if self.sampleinfo:
            print('\033[1;32m校正标准名称\033[0m')
            if warn_novoid:
                print('\033[1;44m>>> 下机单和信息搜集表中名称不一致的样本如下：\033[0m')
                for novoid in warn_novoid:
                    print(novoid)
            if wrong_novoid:
                print('\033[1;44m>>> novoid不在信息搜集表中\033[0m')
                for novoid in wrong_novoid:
                    print(novoid)
            if not warn_novoid and not wrong_novoid:
                print('\033[1;33m全部样本均有标准名称\033[0m')
        else:
            print('\033[1;33m没有进行校正标准名称\033[0m')

    def start(self):
        self.x2t()
        self.get_target_sam_lists() 

        if self.people:
            print('\033[1;32m打印指定信息人员的加测信息: {self.people}\033[0m'.format(**locals()))
            self.show_people_info()

        if self.number:
            print('\033[1;33m提取指定分期下机路径：{self.number}'.format(**locals()))
            self.extract_target_path()
            print('\033[1;32m结果文件: {self.result}\033[0m'.format(**locals()))

        

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='\033[1;44m处理指定分期的下机单\npython2\033[0m')
    parser.add_argument('infile', help='下机单')
    parser.add_argument('--people', help='展示信息人员的该下机单有哪些加测分期',default='吕梦婷')
    parser.add_argument('--number', help='提取指定分期编号路径，并生成sample_list文件')
    parser.add_argument('--sampleinfo', help='用于校正样本标准名称,信息搜集表')
    parser.add_argument('--result',help='输出samplelist的文件名称')
    parser.add_argument('--targetsam', help='提取指定分期指定样本的的加测单')

    args = vars(parser.parse_args())

    dd = ExtractPath(args)
    dd.start()

