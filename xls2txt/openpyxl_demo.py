#!/usr/bin/env python
#-*- coding:utf-8 -*-


import openpyxl 
from openpyxl import load_workbook    #读取excel文件
import sys

infile = sys.argv[1]

wb = load_workbook(infile)
ws = wb.active


n = 15
while ws['K{}'.format(n)].value:  
    print(ws['K{}'.format(n)].value)
    n += 1

