#!/usr/bin/env python
#-*- coding:utf-8 -*-

import openpyxl 
from openpyxl import load_workbook  #读取excel文件
from openpyxl import Workbook       # 创建xlsx文件


class ChangeType(object):

    def __init__(self, args):

        self.source = args["source"]
        self.target = args["target"]
    
    def xls2txt(self):
        wb = load_workbook(self.source)
        ws = wb.active
        out = open(self.target, 'w')
        for row in  ws.rows:
            for cell in row:
                value = cell.value
                out.write("{value}\t".format(**locals()))
            out.write("\n")

    def txt2xls(self):
        wb = Workbook()
        ws = wb.create_sheet("Result", index=0) 
        with open(self.source, 'r') as fr:
            for line in fr:
                linelist = line.strip().split("\t")
                ws.append(linelist)
        wb.save(self.target)


def main():

    demo = ChangeType(args)

    if args["type"]=="x2t":
        demo.xls2txt()
    elif args["type"] == "t2x":
        demo.txt2xls()
    else:
        exit("请输入想要转换的格式")


if __name__ == "__main__":
    import argparse 
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", help="原格式文件")
    parser.add_argument("--target", help="想要转换的格式文件")
    parser.add_argument("--type", help="请输入想要转换的格式", choices=["x2t", "t2x"])

    args = vars(parser.parse_args())

    main()


